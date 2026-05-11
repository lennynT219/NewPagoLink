from typing import Any, Dict, Optional
import csv
import io
import logging

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Sum
from decimal import Decimal
from django.db import transaction
from django.conf import settings
from payments.models import Link, Payment, Refund, Sale
from accounts.models import CustomUser
from shared.email_service import send_html_email

logger = logging.getLogger(__name__)


def calculate_tax(amount: Decimal, include_tax: bool) -> Decimal:
  """Calcula el IVA (dinámico desde settings) si no está incluido."""
  if not include_tax:
    return (amount * settings.TAX_IVA).quantize(Decimal('0.01'))
  return Decimal('0.00')


def send_payment_invite(payment: Payment, req: Any) -> None:
  """Envía un correo al cliente invitándolo a pagar el link generado."""
  if not payment.email:
    return

  subject = f'Invitación de pago de {payment.seller.user.get_full_name()}'
  context = {'payment': payment, 'pay_url': f'{req.scheme}://{req.get_host()}/payments/pay/{payment.link.id}/?p={payment.id}'}
  ok = send_html_email(subject, [payment.email], 'payments/emails/invite_email.html', context)
  if not ok:
    logger.error('Payment invite email failed for %s', payment.email)


def send_payment_confirmation(payment: Payment) -> None:
  """Envía el comprobante de pago exitoso al cliente."""
  if not payment.email:
    return

  subject = f'Comprobante de Pago Exitoso - {payment.description}'
  ok = send_html_email(subject, [payment.email], 'payments/emails/confirmation_email.html', {'payment': payment})
  if not ok:
    logger.error('Payment confirmation email failed for %s', payment.email)


@transaction.atomic
def create_payment_link(seller: CustomUser, data: Dict[str, Any]) -> Link:
  """Servicio para crear un Link de pago con lógica financiera avanzada (IVA 15%/0%)."""
  input_amount = Decimal(str(data['subtotal']))
  tax_type = data.get('tax_type', Link.TaxType.VAT_15)
  include_igv = data.get('include_igv', True)

  # Determinar tasa de IVA
  if tax_type == Link.TaxType.VAT_15:
    vat_rate = Decimal('0.15')
  elif tax_type == Link.TaxType.VAT_0:
    vat_rate = Decimal('0.00')
  else:
    vat_rate = Decimal('0.00')

  if include_igv:
    # El monto ingresado ya tiene IVA. Desglosamos hacia atrás.
    # Total = Monto. Base = Total / (1 + tasa)
    total_amount = input_amount
    base_imponible = (total_amount / (Decimal('1.00') + vat_rate)).quantize(Decimal('0.01'))
    valor_iva = total_amount - base_imponible
  else:
    # El monto ingresado es la base. Sumamos IVA.
    base_imponible = input_amount
    valor_iva = (base_imponible * vat_rate).quantize(Decimal('0.01'))
    total_amount = base_imponible + valor_iva

  return Link.objects.create(
    seller=seller,
    description=data['description'],
    tax_type=tax_type,
    vat_rate=vat_rate,
    include_igv=include_igv,
    subtotal=base_imponible,
    igv=valor_iva,
    amount=total_amount,
    unique=data.get('unique', False),
  )


@transaction.atomic
def process_payment_result(payment_id: int, result_data: Dict[str, Any]) -> Payment:
  """Actualiza el estado del pago y gestiona links únicos y contabilidad."""
  payment = Payment.objects.select_for_update().get(id=payment_id)
  result_code = result_data.get('result', {}).get('code', '')
  SUCCESS_CODES = ['000.000.000', '000.100.110', '000.100.112', '000.400.010', '000.400.020']

  if any(code in result_code for code in SUCCESS_CODES):
    payment.status = Payment.PaymentStatus.PAID
    payment.state = True
    payment.transaction_id = result_data.get('id')
    if payment.link and payment.link.unique:
      payment.link.active = False
      payment.link.save()

    # Registro contable de la venta
    if not hasattr(payment, 'sale_record'):
      net_amount = payment.amount - payment.commission
      Sale.objects.create(
        vendor=payment.seller,
        payment=payment,
        amount=payment.amount,
        commission=payment.commission,
        net_amount=net_amount,
      )

      # Actualizar saldo del vendedor
      if payment.seller:
        seller = CustomUser.objects.select_for_update().get(id=payment.seller.id)
        seller.available_balance += net_amount
        seller.save()

    # Notificar al cliente
    try:
      send_payment_confirmation(payment)
    except Exception as e:
      logger.error('Error sending payment confirmation for payment %s: %s', payment_id, e)
  else:
    payment.status = Payment.PaymentStatus.FAILED
    payment.state = False

  payment.save()
  return payment


def generate_payments_csv(seller: CustomUser) -> io.StringIO:
  """Genera reporte en CSV."""
  payments = Payment.objects.filter(seller=seller, state=True).order_by('-id')
  output = io.StringIO()
  writer = csv.writer(output)
  writer.writerow(['FECHA', 'CLIENTE', 'IDENTIFICACIÓN', 'CORREO', 'TOTAL', 'ID TRANSACCIÓN'])
  for p in payments:
    writer.writerow(
      [
        p.link.created_at.strftime('%Y-%m-%d') if p.link else '',
        f'{p.first_name} {p.last_name}',
        p.identify,
        p.email,
        p.amount,
        p.transaction_id,
      ]
    )
  return output


def generate_payments_excel(seller: CustomUser) -> io.BytesIO:
  """Genera reporte en Excel (.xlsx) con estilos."""
  payments = Payment.objects.filter(seller=seller, state=True).order_by('-id')
  wb = Workbook()
  ws = wb.active
  ws.title = 'Ventas'
  headers = ['FECHA', 'CLIENTE', 'IDENTIFICACIÓN', 'CORREO', 'TOTAL', 'ID TRANSACCIÓN']
  ws.append(headers)

  for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')

  for p in payments:
    ws.append(
      [
        p.link.created_at.strftime('%Y-%m-%d') if p.link else '',
        f'{p.first_name} {p.last_name}',
        p.identify,
        p.email,
        float(p.amount),
        p.transaction_id,
      ]
    )

  output = io.BytesIO()
  wb.save(output)
  output.seek(0)
  return output


@transaction.atomic
def request_refund(seller: CustomUser, payment: Payment, description: str) -> Refund:
  """Crea una solicitud de reembolso para un pago específico."""
  if payment.seller != seller:
    raise ValueError('No tiene permisos sobre este pago.')
  if payment.status != Payment.PaymentStatus.PAID:
    raise ValueError('Solo se pueden reembolsar pagos confirmados.')
  if hasattr(payment, 'refund'):
    raise ValueError('Este pago ya tiene una solicitud de reembolso.')

  return Refund.objects.create(
    seller=seller, payment=payment, description=description, amount=payment.amount, state=False
  )


from django.utils import timezone


@transaction.atomic
def approve_refund_logic(refund_id: int, admin_user: Any) -> Refund:
  """
  Lógica de negocio para aprobar un reembolso.
  1. Marca el Refund como aprobado.
  2. Actualiza el Payment relacionado a un estado de 'Reembolsado' (via state=False o status).
  3. Registra auditoría (quién y cuándo).
  """
  refund = Refund.objects.select_for_update().get(id=refund_id)
  if refund.status != Refund.RefundStatus.PENDING:
    raise ValueError("Este reembolso ya ha sido procesado anteriormente.")

  # Actualizar Reembolso
  refund.status = Refund.RefundStatus.APPROVED
  refund.processed_by = admin_user
  refund.processed_at = timezone.now()
  refund.save()

  # Actualizar Pago relacionado
  payment = refund.payment
  payment.status = Payment.PaymentStatus.FAILED # O añadir un estado 'REFUNDED' si se desea
  payment.state = False # Ya no cuenta como venta activa para KPIs
  payment.save()

  return refund



def get_seller_stats(seller: CustomUser) -> Dict[str, Any]:
  """Estadísticas del Dashboard incluyendo KPIs financieros."""
  today = timezone.now().date()

  links_count = Link.objects.filter(seller=seller).count()

  # Ventas totales históricas (Pagadas)
  sales_agg = Payment.objects.filter(seller=seller, state=True).aggregate(total=Sum('amount'))
  total_sales = sales_agg['total'] or Decimal('0.00')

  # Ventas del día
  daily_sales_agg = Payment.objects.filter(seller=seller, state=True, link__created_at__date=today).aggregate(
    total=Sum('amount')
  )
  daily_sales = daily_sales_agg['total'] or Decimal('0.00')

  # Reembolsos totales
  refunds_agg = Refund.objects.filter(seller=seller, status=Refund.RefundStatus.APPROVED).aggregate(total=Sum('amount'))
  total_refunds = refunds_agg['total'] or Decimal('0.00')

  return {
    'links_count': links_count,
    'total_sales': total_sales,
    'daily_sales': daily_sales,
    'total_refunds': total_refunds,
    'available_balance': seller.available_balance,
    'active': seller.state,
    'email_active': seller.email_active,
    'seller_name': seller.user.first_name,
  }
